from fastapi import FastAPI, Request, HTTPException, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlmodel import SQLModel, Session, create_engine, select
from typing import Optional
from utils.print_receipt import do_print_receipt
from models import GoldTestingTransactions
from datetime import datetime
from pathlib import Path
from collections import defaultdict
import io
import calendar

# Get the path to the directory of this main.py file
BASE_DIR = Path(__file__).parent

# Place your database in the backend directory
DB_PATH = BASE_DIR /"db"/ "GK_DB.db"
DATABASE_URL = f"sqlite:///{DB_PATH}"

engine = create_engine(DATABASE_URL, echo=True)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
def on_startup():
    SQLModel.metadata.create_all(engine)
    print("Database and tables created")

# ---- Request Model ----
class TransactionCreate(SQLModel):
    CustomerName: str
    CustomerMobile: str
    SampleWeight: float
    SampleType: str
    TouchValue: float
    KaratValue: float
    TestingMethod: str                    # <--- ADDED
    Remark: Optional[str] = ""

# ---- CREATE ----
@app.post("/entries")
def create_entry(entry: TransactionCreate = Body(...)):
    txn = GoldTestingTransactions(
        CustomerName=entry.CustomerName,
        CustomerMobile=entry.CustomerMobile,
        SampleWeight=entry.SampleWeight,
        SampleType=entry.SampleType,
        TouchValue=entry.TouchValue,
        KaratValue=entry.KaratValue,
        TestingMethod=entry.TestingMethod,  # <--- ADDED
        Remark=entry.Remark,
        TransactionDate=datetime.now().isoformat(),
        TestedOn=None
    )
    with Session(engine) as session:
        session.add(txn)
        session.commit()
        session.refresh(txn)
        return {
            "TransactionID": txn.TransactionID,
            "TransactionDate": txn.TransactionDate,
            "TestedOn": txn.TestedOn,
            "CustomerName": txn.CustomerName,
            "CustomerMobile": txn.CustomerMobile,
            "SampleWeight": txn.SampleWeight,
            "SampleType": txn.SampleType,
            "TouchValue": txn.TouchValue,
            "KaratValue": txn.KaratValue,
            "TestingMethod": txn.TestingMethod,  # <--- ADDED
            "Remark": txn.Remark
        }

# ---- READ ALL ----
@app.get("/entries")
def get_entries():
    with Session(engine) as session:
        txns = session.exec(select(GoldTestingTransactions)).all()
        result = []
        for t in txns:
            result.append({
                "TransactionID": t.TransactionID,
                "TransactionDate": t.TransactionDate,
                "TestedOn": t.TestedOn,
                "CustomerName": t.CustomerName,
                "CustomerMobile": t.CustomerMobile,
                "SampleWeight": t.SampleWeight,
                "SampleType": t.SampleType,
                "TouchValue": t.TouchValue,
                "KaratValue": t.KaratValue,
                "TestingMethod": t.TestingMethod,  # <--- ADDED
                "Remark": t.Remark
            })
        return result

# ---- READ SINGLE ----
@app.get("/entries/{entry_id}")
def get_entry(entry_id: int):
    with Session(engine) as session:
        txn = session.get(GoldTestingTransactions, entry_id)
        if txn is None:
            raise HTTPException(status_code=404, detail="GoldTestingTransactions not found")
        return {
            "TransactionID": txn.TransactionID,
            "TransactionDate": txn.TransactionDate,
            "TestedOn": txn.TestedOn,
            "CustomerName": txn.CustomerName,
            "CustomerMobile": txn.CustomerMobile,
            "SampleWeight": txn.SampleWeight,
            "SampleType": txn.SampleType,
            "TouchValue": txn.TouchValue,
            "KaratValue": txn.KaratValue,
            "TestingMethod": txn.TestingMethod,  # <--- ADDED
            "Remark": txn.Remark
        }

# ---- UPDATE ----
@app.put("/entries/{entry_id}")
def update_entry(entry_id: int, entry: TransactionCreate = Body(...)):
    with Session(engine) as session:
        txn = session.get(GoldTestingTransactions, entry_id)
        if txn is None:
            raise HTTPException(status_code=404, detail="GoldTestingTransactions not found")
        txn.CustomerName = entry.CustomerName
        txn.CustomerMobile = entry.CustomerMobile
        txn.SampleWeight = entry.SampleWeight
        txn.SampleType = entry.SampleType
        txn.TouchValue = entry.TouchValue
        txn.KaratValue = entry.KaratValue
        txn.TestingMethod = entry.TestingMethod  # <--- ADDED
        txn.Remark = entry.Remark
        txn.TestedOn = datetime.now().isoformat()
        session.add(txn)
        session.commit()
        session.refresh(txn)
        return {
            "TransactionID": txn.TransactionID,
            "TransactionDate": txn.TransactionDate,
            "TestedOn": txn.TestedOn,
            "CustomerName": txn.CustomerName,
            "CustomerMobile": txn.CustomerMobile,
            "SampleWeight": txn.SampleWeight,
            "SampleType": txn.SampleType,
            "TouchValue": txn.TouchValue,
            "KaratValue": txn.KaratValue,
            "TestingMethod": txn.TestingMethod,  # <--- ADDED
            "Remark": txn.Remark
        }

# ---- DELETE ----
@app.delete("/entries/{entry_id}")
def delete_entry(entry_id: int):
    with Session(engine) as session:
        txn = session.get(GoldTestingTransactions, entry_id)
        if txn is None:
            raise HTTPException(status_code=404, detail="GoldTestingTransactions not found")
        session.delete(txn)
        session.commit()
        return {"detail": "GoldTestingTransactions deleted", "TransactionID": entry_id}

# ---------- PRINT RECEIPT ----
@app.post("/print-receipt")
async def print_receipt(request: Request):
    data = await request.json()
    copies = int(data.get("Copies", 1))
    entry = data.get("Entry", {})
    print("Printing receipt with data:", entry, "Copies:", copies)
    result = do_print_receipt(entry, copies)
    return result


# ---------- VOICE RECOGNITION ----
@app.post("/voice-recognize")
async def voice_recognize(request: Request):
    """Receive WAV audio and return recognized text using Google Speech API."""
    import tempfile
    import os
    try:
        import speech_recognition as sr
    except ImportError:
        raise HTTPException(status_code=500, detail="SpeechRecognition not installed. Run: pip install SpeechRecognition")

    audio_bytes = await request.body()
    if not audio_bytes or len(audio_bytes) < 100:
        return {"text": "", "alternatives": [], "error": "No audio data or too short"}

    # Save to temp WAV file
    tmp_path = None
    try:
        fd, tmp_path = tempfile.mkstemp(suffix=".wav")
        os.write(fd, audio_bytes)
        os.close(fd)

        recognizer = sr.Recognizer()
        recognizer.energy_threshold = 200
        recognizer.dynamic_energy_threshold = True

        with sr.AudioFile(tmp_path) as source:
            # Skip first 0.3s of possible silence/click
            recognizer.adjust_for_ambient_noise(source, duration=0.3)
            audio = recognizer.record(source)

        # Google (free, best for Indian English names)
        results = recognizer.recognize_google(audio, language="en-IN", show_all=True)

        # Handle various response formats from Google
        if not results:
            return {"text": "", "alternatives": [], "error": "Could not understand audio"}

        alternatives = []
        if isinstance(results, dict):
            # results = {"alternative": [{"transcript": "...", "confidence": ...}, ...]}
            alts = results.get("alternative", [])
            if not alts:
                # Sometimes nested under 'results'
                for r in results.get("results", []):
                    alts.extend(r.get("alternatives", []))
            alternatives = [a.get("transcript", "") for a in alts[:5]]
        elif isinstance(results, list):
            # Sometimes returns a list of dicts
            for r in results:
                if isinstance(r, dict):
                    for a in r.get("alternative", []):
                        alternatives.append(a.get("transcript", ""))
                elif isinstance(r, str):
                    alternatives.append(r)

        if not alternatives:
            return {"text": "", "alternatives": [], "error": "Could not understand audio"}

        best = alternatives[0]
        return {"text": best.upper(), "alternatives": [a.upper() for a in alternatives[:5]]}

    except sr.UnknownValueError:
        return {"text": "", "alternatives": [], "error": "Could not understand audio"}
    except sr.RequestError as e:
        return {"text": "", "alternatives": [], "error": f"Google API error: {str(e)}"}
    except Exception as e:
        return {"text": "", "alternatives": [], "error": f"Processing error: {str(e)}"}
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except:
                pass


# ---------- REPORT / DASHBOARD ----
@app.get("/report/summary")
def get_report_summary():
    """Returns dashboard data: today stats, monthly summary, customer analytics."""
    with Session(engine) as session:
        txns = session.exec(select(GoldTestingTransactions)).all()

    today_str = datetime.now().strftime("%Y-%m-%d")
    today_entries = [t for t in txns if t.TransactionDate and t.TransactionDate.startswith(today_str)]

    # Monthly summary
    monthly = defaultdict(lambda: {"entries": 0, "customers": set(), "weight": 0.0, "gold_count": 0, "gold_weight": 0.0, "silver_count": 0, "silver_weight": 0.0})
    for t in txns:
        if not t.TransactionDate:
            continue
        month_key = t.TransactionDate[:7]  # "2026-05"
        monthly[month_key]["entries"] += 1
        monthly[month_key]["customers"].add((t.CustomerName or "").upper())
        monthly[month_key]["weight"] += t.SampleWeight or 0
        st = (t.SampleType or "").upper()
        if "GOLD" in st:
            monthly[month_key]["gold_count"] += 1
            monthly[month_key]["gold_weight"] += t.SampleWeight or 0
        elif "SILVER" in st:
            monthly[month_key]["silver_count"] += 1
            monthly[month_key]["silver_weight"] += t.SampleWeight or 0

    monthly_list = []
    for k in sorted(monthly.keys(), reverse=True):
        v = monthly[k]
        monthly_list.append({
            "month": k,
            "entries": v["entries"],
            "customers": len(v["customers"]),
            "weight": round(v["weight"], 3),
            "gold_count": v["gold_count"],
            "gold_weight": round(v["gold_weight"], 3),
            "silver_count": v["silver_count"],
            "silver_weight": round(v["silver_weight"], 3),
        })

    # Customer analytics
    cust_data = defaultdict(lambda: {"mobile": "", "visits": 0, "last_visit": "", "touch_sum": 0.0, "touch_count": 0, "samples": defaultdict(int)})
    for t in txns:
        name = (t.CustomerName or "").upper()
        if not name:
            continue
        c = cust_data[name]
        c["visits"] += 1
        if t.CustomerMobile:
            c["mobile"] = t.CustomerMobile
        if t.TransactionDate and t.TransactionDate > c["last_visit"]:
            c["last_visit"] = t.TransactionDate
        if t.TouchValue and t.TouchValue > 0:
            c["touch_sum"] += t.TouchValue
            c["touch_count"] += 1
        if t.SampleType:
            c["samples"][t.SampleType] += 1

    customer_list = []
    for name, c in sorted(cust_data.items(), key=lambda x: -x[1]["visits"]):
        most_common = max(c["samples"], key=c["samples"].get) if c["samples"] else ""
        avg_touch = round(c["touch_sum"] / c["touch_count"], 2) if c["touch_count"] > 0 else 0
        days_since = 0
        if c["last_visit"]:
            try:
                last_dt = datetime.fromisoformat(c["last_visit"])
                days_since = (datetime.now() - last_dt).days
            except:
                pass
        customer_list.append({
            "name": name,
            "mobile": c["mobile"],
            "visits": c["visits"],
            "last_visit": c["last_visit"][:10] if c["last_visit"] else "",
            "avg_touch": avg_touch,
            "most_common_sample": most_common,
            "days_since_last": days_since,
        })

    # Sample Type Breakdown
    type_data = defaultdict(lambda: {"count": 0, "touch_sum": 0.0, "touch_count": 0, "weight_sum": 0.0})
    total = len(txns)
    for t in txns:
        st = t.SampleType or "Unknown"
        type_data[st]["count"] += 1
        type_data[st]["weight_sum"] += t.SampleWeight or 0
        if t.TouchValue and t.TouchValue > 0:
            type_data[st]["touch_sum"] += t.TouchValue
            type_data[st]["touch_count"] += 1
    sample_types = []
    for st, v in sorted(type_data.items(), key=lambda x: -x[1]["count"]):
        avg_t = round(v["touch_sum"] / v["touch_count"], 2) if v["touch_count"] > 0 else 0
        avg_w = round(v["weight_sum"] / v["count"], 3) if v["count"] > 0 else 0
        pct = round(v["count"] / total * 100, 1) if total > 0 else 0
        sample_types.append({"type": st, "count": v["count"], "avg_touch": avg_t, "avg_weight": avg_w, "pct": pct})

    # Peak Hours
    hour_data = defaultdict(int)
    for t in txns:
        if t.TransactionDate and len(t.TransactionDate) >= 13:
            try:
                h = int(t.TransactionDate[11:13])
                hour_data[h] += 1
            except:
                pass
    peak_hours = []
    for h in range(7, 22):
        count = hour_data.get(h, 0)
        pct = round(count / total * 100, 1) if total > 0 else 0
        peak_hours.append({"hour": f"{h:02d}:00-{h+1:02d}:00", "count": count, "pct": pct})

    # Day-of-Week distribution
    dow_data = defaultdict(int)
    for t in txns:
        if t.TransactionDate and len(t.TransactionDate) >= 10:
            try:
                dt = datetime.fromisoformat(t.TransactionDate[:10])
                dow_data[dt.weekday()] += 1
            except:
                pass
    day_of_week = []
    for d in range(7):  # 0=Monday to 6=Sunday
        count = dow_data.get(d, 0)
        pct = round(count / total * 100, 1) if total > 0 else 0
        day_of_week.append({"day": calendar.day_name[d], "count": count, "pct": pct})

    # Customer Churn: customers with 2+ visits who haven't returned in 30+ days
    churn_list = []
    for name, c in cust_data.items():
        if c["visits"] >= 2 and c["last_visit"]:
            try:
                last_dt = datetime.fromisoformat(c["last_visit"])
                days_gone = (datetime.now() - last_dt).days
                if days_gone >= 30:
                    avg_touch = round(c["touch_sum"] / c["touch_count"], 2) if c["touch_count"] > 0 else 0
                    most_common = max(c["samples"], key=c["samples"].get) if c["samples"] else ""
                    churn_list.append({
                        "name": name,
                        "mobile": c["mobile"],
                        "visits": c["visits"],
                        "last_visit": c["last_visit"][:10],
                        "days_gone": days_gone,
                        "avg_touch": avg_touch,
                        "most_common_sample": most_common,
                    })
            except:
                pass
    churn_list.sort(key=lambda x: (-x["visits"], -x["days_gone"]))

    # New vs Returning per month
    customer_first_visit = {}
    for t in sorted(txns, key=lambda x: x.TransactionDate or ""):
        name = (t.CustomerName or "").upper()
        if not name or not t.TransactionDate:
            continue
        if name not in customer_first_visit:
            customer_first_visit[name] = t.TransactionDate[:7]
    new_vs_returning = {}
    for t in txns:
        if not t.TransactionDate:
            continue
        m = t.TransactionDate[:7]
        name = (t.CustomerName or "").upper()
        if not name:
            continue
        if m not in new_vs_returning:
            new_vs_returning[m] = {"new": set(), "returning": set()}
        if customer_first_visit.get(name) == m:
            new_vs_returning[m]["new"].add(name)
        else:
            new_vs_returning[m]["returning"].add(name)
    new_ret_list = []
    for m in sorted(new_vs_returning.keys(), reverse=True):
        nv = new_vs_returning[m]
        new_count = len(nv["new"])
        ret_count = len(nv["returning"])
        total_m = new_count + ret_count
        new_ret_list.append({
            "month": m,
            "new": new_count,
            "returning": ret_count,
            "total": total_m,
            "new_pct": round(new_count / total_m * 100, 1) if total_m > 0 else 0,
        })

    # Customer Loyalty Segments
    segments = {"one_time": 0, "occasional": 0, "regular": 0, "vip": 0}
    for name, c in cust_data.items():
        v = c["visits"]
        if v == 1:
            segments["one_time"] += 1
        elif v <= 3:
            segments["occasional"] += 1
        elif v <= 9:
            segments["regular"] += 1
        else:
            segments["vip"] += 1
    total_customers = sum(segments.values())
    loyalty_segments = [
        {"segment": "One-Time (1 visit)", "count": segments["one_time"], "pct": round(segments["one_time"] / total_customers * 100, 1) if total_customers else 0},
        {"segment": "Occasional (2-3 visits)", "count": segments["occasional"], "pct": round(segments["occasional"] / total_customers * 100, 1) if total_customers else 0},
        {"segment": "Regular (4-9 visits)", "count": segments["regular"], "pct": round(segments["regular"] / total_customers * 100, 1) if total_customers else 0},
        {"segment": "VIP (10+ visits)", "count": segments["vip"], "pct": round(segments["vip"] / total_customers * 100, 1) if total_customers else 0},
    ]

    # Month-over-Month Growth
    growth_list = []
    sorted_months = sorted(monthly_list, key=lambda x: x["month"])
    for i in range(1, len(sorted_months)):
        prev = sorted_months[i - 1]
        curr = sorted_months[i]
        entry_growth = round((curr["entries"] - prev["entries"]) / prev["entries"] * 100, 1) if prev["entries"] > 0 else 0
        cust_growth = round((curr["customers"] - prev["customers"]) / prev["customers"] * 100, 1) if prev["customers"] > 0 else 0
        growth_list.append({
            "month": curr["month"],
            "entries": curr["entries"],
            "prev_entries": prev["entries"],
            "entry_growth": entry_growth,
            "customers": curr["customers"],
            "prev_customers": prev["customers"],
            "cust_growth": cust_growth,
        })
    growth_list.reverse()

    # Touch Value Distribution (purity ranges)
    touch_ranges = {"90-100 (High)": 0, "80-90 (Good)": 0, "70-80 (Medium)": 0, "60-70 (Fair)": 0, "Below 60": 0, "Not Tested": 0}
    for t in txns:
        tv = t.TouchValue or 0
        if tv == 0:
            touch_ranges["Not Tested"] += 1
        elif tv >= 90:
            touch_ranges["90-100 (High)"] += 1
        elif tv >= 80:
            touch_ranges["80-90 (Good)"] += 1
        elif tv >= 70:
            touch_ranges["70-80 (Medium)"] += 1
        elif tv >= 60:
            touch_ranges["60-70 (Fair)"] += 1
        else:
            touch_ranges["Below 60"] += 1
    touch_distribution = []
    for label, count in touch_ranges.items():
        if count > 0:
            touch_distribution.append({"range": label, "count": count, "pct": round(count / total * 100, 1) if total > 0 else 0})

    # Average Test Time (entries that have both TransactionDate and TestedOn)
    test_times = []
    for t in txns:
        if t.TransactionDate and t.TestedOn:
            try:
                created = datetime.fromisoformat(t.TransactionDate)
                tested = datetime.fromisoformat(t.TestedOn)
                diff_min = (tested - created).total_seconds() / 60
                if 0 < diff_min < 1440:  # within 24h
                    test_times.append(diff_min)
            except:
                pass
    avg_test_min = round(sum(test_times) / len(test_times), 1) if test_times else 0

    return {
        "today": {
            "entries": len(today_entries),
            "customers": len(set((t.CustomerName or "").upper() for t in today_entries if t.CustomerName)),
            "weight": round(sum(t.SampleWeight or 0 for t in today_entries), 3),
        },
        "monthly": monthly_list,
        "customers": customer_list[:50],
        "sample_types": sample_types,
        "peak_hours": peak_hours,
        "day_of_week": day_of_week,
        "churn_customers": churn_list[:30],
        "new_vs_returning": new_ret_list,
        "loyalty_segments": loyalty_segments,
        "growth": growth_list,
        "touch_distribution": touch_distribution,
        "avg_test_min": avg_test_min,
        "total_entries": total,
        "total_customers": total_customers,
    }


@app.get("/report/download")
def download_report():
    """Generate and return multi-sheet Excel report."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    with Session(engine) as session:
        txns = session.exec(select(GoldTestingTransactions)).all()

    wb = Workbook()

    # --- Sheet 1: Daily Summary ---
    ws = wb.active
    ws.title = "Daily Summary"
    ws.append(["Date", "Entries", "Customers", "Total Weight (gm)"])
    daily = defaultdict(lambda: {"entries": 0, "customers": set(), "weight": 0.0})
    for t in txns:
        if not t.TransactionDate:
            continue
        d = t.TransactionDate[:10]
        daily[d]["entries"] += 1
        daily[d]["customers"].add((t.CustomerName or "").upper())
        daily[d]["weight"] += t.SampleWeight or 0
    for d in sorted(daily.keys(), reverse=True):
        ws.append([d, daily[d]["entries"], len(daily[d]["customers"]), round(daily[d]["weight"], 3)])

    # --- Sheet 2: Monthly Summary ---
    ws2 = wb.create_sheet("Monthly Summary")
    ws2.append(["Month", "Entries", "Customers", "Total Weight (gm)"])
    monthly = defaultdict(lambda: {"entries": 0, "customers": set(), "weight": 0.0})
    for t in txns:
        if not t.TransactionDate:
            continue
        m = t.TransactionDate[:7]
        monthly[m]["entries"] += 1
        monthly[m]["customers"].add((t.CustomerName or "").upper())
        monthly[m]["weight"] += t.SampleWeight or 0
    for m in sorted(monthly.keys(), reverse=True):
        ws2.append([m, monthly[m]["entries"], len(monthly[m]["customers"]), round(monthly[m]["weight"], 3)])

    # --- Sheet 3: Monthly Gold vs Silver ---
    ws3 = wb.create_sheet("Gold vs Silver Monthly")
    ws3.append(["Month", "Gold Count", "Gold Weight (gm)", "Silver Count", "Silver Weight (gm)", "Other Count", "Other Weight (gm)"])
    gs_monthly = defaultdict(lambda: {"gc": 0, "gw": 0.0, "sc": 0, "sw": 0.0, "oc": 0, "ow": 0.0})
    for t in txns:
        if not t.TransactionDate:
            continue
        m = t.TransactionDate[:7]
        st = (t.SampleType or "").upper()
        w = t.SampleWeight or 0
        if "GOLD" in st:
            gs_monthly[m]["gc"] += 1
            gs_monthly[m]["gw"] += w
        elif "SILVER" in st:
            gs_monthly[m]["sc"] += 1
            gs_monthly[m]["sw"] += w
        else:
            gs_monthly[m]["oc"] += 1
            gs_monthly[m]["ow"] += w
    for m in sorted(gs_monthly.keys(), reverse=True):
        v = gs_monthly[m]
        ws3.append([m, v["gc"], round(v["gw"], 3), v["sc"], round(v["sw"], 3), v["oc"], round(v["ow"], 3)])

    # --- Sheet 4: Customer Analytics ---
    ws4 = wb.create_sheet("Customer Analytics")
    ws4.append(["Customer", "Mobile", "Visits", "Last Visit", "Avg Touch", "Most Common Sample", "Days Since Last"])
    cust_data = defaultdict(lambda: {"mobile": "", "visits": 0, "last_visit": "", "touch_sum": 0.0, "touch_count": 0, "samples": defaultdict(int)})
    for t in txns:
        name = (t.CustomerName or "").upper()
        if not name:
            continue
        c = cust_data[name]
        c["visits"] += 1
        if t.CustomerMobile:
            c["mobile"] = t.CustomerMobile
        if t.TransactionDate and t.TransactionDate > c["last_visit"]:
            c["last_visit"] = t.TransactionDate
        if t.TouchValue and t.TouchValue > 0:
            c["touch_sum"] += t.TouchValue
            c["touch_count"] += 1
        if t.SampleType:
            c["samples"][t.SampleType] += 1
    for name, c in sorted(cust_data.items(), key=lambda x: -x[1]["visits"]):
        most_common = max(c["samples"], key=c["samples"].get) if c["samples"] else ""
        avg_touch = round(c["touch_sum"] / c["touch_count"], 2) if c["touch_count"] > 0 else 0
        days_since = 0
        if c["last_visit"]:
            try:
                days_since = (datetime.now() - datetime.fromisoformat(c["last_visit"])).days
            except:
                pass
        ws4.append([name, c["mobile"], c["visits"], c["last_visit"][:10] if c["last_visit"] else "", avg_touch, most_common, days_since])

    # --- Sheet 5: Sample Type Breakdown ---
    ws5 = wb.create_sheet("Sample Type Breakdown")
    ws5.append(["Sample Type", "Count", "Avg Touch", "Avg Weight (gm)", "% of Total"])
    type_data = defaultdict(lambda: {"count": 0, "touch_sum": 0.0, "touch_count": 0, "weight_sum": 0.0})
    total = len(txns)
    for t in txns:
        st = t.SampleType or "Unknown"
        type_data[st]["count"] += 1
        type_data[st]["weight_sum"] += t.SampleWeight or 0
        if t.TouchValue and t.TouchValue > 0:
            type_data[st]["touch_sum"] += t.TouchValue
            type_data[st]["touch_count"] += 1
    for st, v in sorted(type_data.items(), key=lambda x: -x[1]["count"]):
        avg_t = round(v["touch_sum"] / v["touch_count"], 2) if v["touch_count"] > 0 else 0
        avg_w = round(v["weight_sum"] / v["count"], 3) if v["count"] > 0 else 0
        pct = round(v["count"] / total * 100, 1) if total > 0 else 0
        ws5.append([st, v["count"], avg_t, avg_w, f"{pct}%"])

    # --- Sheet 6: Peak Hours ---
    ws6 = wb.create_sheet("Peak Hours")
    ws6.append(["Hour", "Entries", "% of Total"])
    hour_data = defaultdict(int)
    for t in txns:
        if t.TransactionDate and len(t.TransactionDate) >= 13:
            try:
                h = int(t.TransactionDate[11:13])
                hour_data[h] += 1
            except:
                pass
    for h in range(7, 22):  # 7 AM to 9 PM
        count = hour_data.get(h, 0)
        pct = round(count / total * 100, 1) if total > 0 else 0
        ws6.append([f"{h:02d}:00 - {h+1:02d}:00", count, f"{pct}%"])

    # --- Sheet 7: Day of Week ---
    ws7 = wb.create_sheet("Day of Week")
    ws7.append(["Day", "Entries", "% of Total"])
    dow_data = defaultdict(int)
    day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    for t in txns:
        if t.TransactionDate and len(t.TransactionDate) >= 10:
            try:
                dt = datetime.fromisoformat(t.TransactionDate[:10])
                dow_data[dt.weekday()] += 1
            except:
                pass
    for i, name in enumerate(day_names):
        count = dow_data.get(i, 0)
        pct = round(count / total * 100, 1) if total > 0 else 0
        ws7.append([name, count, f"{pct}%"])

    # --- Sheet 8: Growth (Month-over-Month) ---
    ws8 = wb.create_sheet("Growth")
    ws8.append(["Month", "Entries", "Prev Entries", "Entry Growth %", "Customers", "Prev Customers", "Cust Growth %"])
    sorted_months = sorted(monthly.keys())
    for i, m in enumerate(sorted_months):
        entries = monthly[m]["entries"]
        customers = len(monthly[m]["customers"])
        if i == 0:
            ws8.append([m, entries, "", "", customers, "", ""])
        else:
            pm = sorted_months[i - 1]
            prev_e = monthly[pm]["entries"]
            prev_c = len(monthly[pm]["customers"])
            eg = round((entries - prev_e) / prev_e * 100, 1) if prev_e > 0 else 0
            cg = round((customers - prev_c) / prev_c * 100, 1) if prev_c > 0 else 0
            ws8.append([m, entries, prev_e, f"{eg}%", customers, prev_c, f"{cg}%"])

    # --- Sheet 9: New vs Returning Customers ---
    ws9 = wb.create_sheet("New vs Returning")
    ws9.append(["Month", "New Customers", "Returning Customers", "Total", "New %"])
    first_seen = {}
    for t in sorted(txns, key=lambda x: x.TransactionDate or ""):
        if not t.TransactionDate or not t.CustomerName:
            continue
        name = (t.CustomerName or "").upper()
        m = t.TransactionDate[:7]
        if name not in first_seen:
            first_seen[name] = m
    nvr_monthly = defaultdict(lambda: {"new": set(), "ret": set()})
    for t in txns:
        if not t.TransactionDate or not t.CustomerName:
            continue
        name = (t.CustomerName or "").upper()
        m = t.TransactionDate[:7]
        if first_seen.get(name) == m:
            nvr_monthly[m]["new"].add(name)
        else:
            nvr_monthly[m]["ret"].add(name)
    for m in sorted(nvr_monthly.keys(), reverse=True):
        new_c = len(nvr_monthly[m]["new"])
        ret_c = len(nvr_monthly[m]["ret"])
        tot = new_c + ret_c
        pct = round(new_c / tot * 100, 1) if tot > 0 else 0
        ws9.append([m, new_c, ret_c, tot, f"{pct}%"])

    # --- Sheet 10: Customer Churn ---
    ws10 = wb.create_sheet("Customer Churn")
    ws10.append(["Customer", "Mobile", "Past Visits", "Last Visit", "Days Gone", "Avg Touch", "Most Common Sample"])
    today = datetime.now()
    for name, c in sorted(cust_data.items(), key=lambda x: -x[1]["visits"]):
        if c["visits"] < 2:
            continue
        days_gone = 0
        if c["last_visit"]:
            try:
                days_gone = (today - datetime.fromisoformat(c["last_visit"])).days
            except:
                continue
        if days_gone < 30:
            continue
        avg_touch = round(c["touch_sum"] / c["touch_count"], 2) if c["touch_count"] > 0 else 0
        most_common = max(c["samples"], key=c["samples"].get) if c["samples"] else ""
        ws10.append([name, c["mobile"], c["visits"], c["last_visit"][:10] if c["last_visit"] else "", days_gone, avg_touch, most_common])

    # --- Sheet 11: Touch Quality Distribution ---
    ws11 = wb.create_sheet("Touch Quality")
    ws11.append(["Range", "Count", "% of Total"])
    range_labels = ["90-100 (Premium)", "80-90 (High)", "70-80 (Medium)", "60-70 (Low)", "Below 60"]
    range_counts = [0, 0, 0, 0, 0]
    for t in txns:
        tv = t.TouchValue or 0
        if tv > 0:
            if tv >= 90:
                range_counts[0] += 1
            elif tv >= 80:
                range_counts[1] += 1
            elif tv >= 70:
                range_counts[2] += 1
            elif tv >= 60:
                range_counts[3] += 1
            else:
                range_counts[4] += 1
    touch_total = sum(range_counts)
    for idx, label in enumerate(range_labels):
        count = range_counts[idx]
        pct = round(count / touch_total * 100, 1) if touch_total > 0 else 0
        ws11.append([label, count, f"{pct}%"])

    # Bold headers in all sheets
    bold = Font(bold=True)
    for sheet in wb.sheetnames:
        for cell in wb[sheet][1]:
            cell.font = bold

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"GK_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
